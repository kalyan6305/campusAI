"""
Study Planner API Router.
"""

import io
import uuid
import logging
from typing import List, Dict, Any
from fastapi import APIRouter, Depends, HTTPException, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.db.base import get_db
from app.models.user import User
from app.models.study_plan import StudentStudyPlan
from app.agents.study_planner_agent import StudyPlannerAgent
from app.utils.dependencies import get_current_user

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/study-planner", tags=["Study Planner"])

# Singleton agent
planner_agent = StudyPlannerAgent()

@router.post("/generate")
async def generate_plan(
    subject: str = Body(...),
    level: str = Body(...),
    days: int = Body(...),
    hours: float = Body(...),
    exam_date: str = Body(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    try:
        plan_data = await planner_agent.generate_plan(subject, level, days, hours, exam_date)
        
        # Create a new session record
        session_id = str(uuid.uuid4())
        new_plan = StudentStudyPlan(
            session_id=session_id,
            user_id=current_user.id,
            subject=subject,
            level=level,
            plan_data=plan_data,
            progress_data={}
        )
        db.add(new_plan)
        await db.commit()
        
        return {"session_id": session_id, "plan": plan_data}
    except Exception as e:
        logger.error(f"Error generating plan: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/list-plans")
async def list_plans(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(StudentStudyPlan).where(
            StudentStudyPlan.user_id == current_user.id
        ).order_by(StudentStudyPlan.created_at.desc())
    )
    plans = result.scalars().all()
    return [
        {
            "session_id": p.session_id,
            "subject": p.subject,
            "level": p.level,
            "created_at": p.created_at.isoformat()
        } for p in plans
    ]

@router.get("/get-plan/{session_id}")
async def get_plan(
    session_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(StudentStudyPlan).where(
            StudentStudyPlan.session_id == session_id,
            StudentStudyPlan.user_id == current_user.id
        )
    )
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    return {"plan": plan.plan_data, "progress": plan.progress_data}

@router.post("/save-progress")
async def save_progress(
    session_id: str = Body(...),
    progress: Dict[str, bool] = Body(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(StudentStudyPlan).where(
            StudentStudyPlan.session_id == session_id,
            StudentStudyPlan.user_id == current_user.id
        )
    )
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    plan.progress_data = progress
    await db.commit()
    return {"status": "success"}

@router.delete("/clear-plan/{session_id}")
async def clear_plan(
    session_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    await db.execute(
        delete(StudentStudyPlan).where(
            StudentStudyPlan.session_id == session_id,
            StudentStudyPlan.user_id == current_user.id
        )
    )
    await db.commit()
    return {"status": "cleared"}

@router.post("/export-excel")
async def export_excel(
    session_id: str = Body(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(StudentStudyPlan).where(
            StudentStudyPlan.session_id == session_id,
            StudentStudyPlan.user_id == current_user.id
        )
    )
    plan_record = result.scalar_one_or_none()
    if not plan_record:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    plan_data = plan_record.plan_data
    
    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Study Plan"
    
    # Header
    ws.append(["Campus AI Study Plan", "", "", "", ""])
    ws.merge_cells('A1:E1')
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    ws.append([f"Subject: {plan_data['subject']}", f"Level: {plan_data['level']}", f"Goal: {plan_data['exam_date'] or 'N/A'}", "", ""])
    ws.append([])
    
    headers = ["Day", "Week", "Topic", "Hours", "Difficulty", "Rationale"]
    ws.append(headers)
    for cell in ws[WS_CURRENT_ROW := ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    for topic in plan_data.get("all_topics", []):
        ws.append([
            topic['day'],
            topic['week'],
            topic['topic'],
            topic['hours'],
            topic['difficulty'],
            topic['rationale']
        ])
    
    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=study_plan_{plan_data['subject'].replace(' ', '_')}.xlsx"}
    )
